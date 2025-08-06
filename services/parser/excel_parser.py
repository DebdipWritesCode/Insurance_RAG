from openpyxl import load_workbook

def extract_text_from_xlsx(file_path: str) -> list[str]:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    chunks = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))

        # Step 1: Identify the header row
        header_row_index = None
        for i, row in enumerate(rows):
            if row and {"Name", "Mobile Number", "Pincode", "Salary"}.issubset(set(str(cell) for cell in row if cell)):
                header_row_index = i
                break

        # Step 2: Extract message-like unstructured content before the table
        for row in rows[:header_row_index]:
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                chunks.append(row_text.strip())

        # Step 3: Extract structured data after the header
        if header_row_index is not None:
            headers = [str(cell) for cell in rows[header_row_index] if cell]
            for row in rows[header_row_index + 1:]:
                if any(row):  # skip empty rows
                    row_text = " | ".join(
                        f"{header}: {cell}" for header, cell in zip(headers, row) if cell is not None
                    )
                    if row_text.strip():
                        chunks.append(row_text.strip())

    print(f"📊 Extracted {len(chunks)} chunks from Excel")
    print(f"🔍 First chunk (100 chars): {repr(chunks[0][:100]) if chunks else 'No chunks'}")

    return chunks
